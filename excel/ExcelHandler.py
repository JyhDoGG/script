#! /usr/bin/env python3

import openpyxl

iplist=openpyxl.load_workbook(filename='iplist.xlsx')
service=openpyxl.load_workbook(filename='service.xlsx')

ws=iplist.active
ws1=service.active

for x in ws['A']:
    project_name=x.value
    print(project_name)
    c=0

    for y in ws1['B']:
        if y.value == project_name:
            c=1
            print(y.row,y.column)
            ws1['L'+str(y.row)].value=ws['B'+str(x.row)].value
            ws1['M'+str(y.row)].value=ws['F'+str(x.row)].value
            ws1['N'+str(y.row)].value=ws['C'+str(x.row)].value
            ws1['O'+str(y.row)].value=ws['E'+str(x.row)].value
            ws1['P'+str(y.row)].value=ws['G'+str(x.row)].value

    if c==0:
        print('not exist')
        r=ws1.max_row+1
        ws1['B'+str(r)].value=ws['A'+str(x.row)].value
        ws1['C'+str(r)].value=ws['A'+str(x.row)].value
        ws1['L'+str(r)].value=ws['B'+str(x.row)].value
        ws1['M'+str(r)].value=ws['F'+str(x.row)].value
        ws1['N'+str(r)].value=ws['C'+str(x.row)].value
        ws1['O'+str(r)].value=ws['E'+str(x.row)].value
        ws1['P'+str(r)].value=ws['G'+str(x.row)].value


service.save('new.xlsx')




